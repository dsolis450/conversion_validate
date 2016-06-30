import pymssql  
import openpyxl
from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border

conn = pymssql.connect(server='ATLASSPSQL1', user='******', password='******', database='SUPPORT31')  
cursor = conn.cursor()  
Tk().withdraw()
filename = askopenfilename()

wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Schedules')
#one sheet at a time - build specifically for assets	
#connect to database	
#open excel file
#function for each column to be validated
#in function: grab db field data and put in list
#iterate through the columns cells, if it fails validation, mark the cell as red, append comment as to why in front


redFill = PatternFill(start_color='FFFF0000',
							end_color='FFFF0000',
							fill_type='solid')
#validate PROPERTYID
cursor.execute("SELECT lgp_external_property_id FROM lgb_property ORDER BY 1;")  
row = cursor.fetchone()
prop_ids = []  
while row:  
    prop_ids.append(row[0])    
    row = cursor.fetchone()
for cellObj in sheet.columns[0]:
	if cellObj.value == None:
		pass
	else:
		if cellObj.value  not in prop_ids:
			sheet[cellObj.coordinate].fill = redFill

#validate REQUESTTYPE
cursor.execute("select lgv_request_type_desc from lgb_requesttype order by 1;")  
row = cursor.fetchone()
reqt = []  
while row:  
    reqt.append(row[0])    
    row = cursor.fetchone()
for cellObj in sheet.columns[1]:
	if cellObj.value == None:
		pass
	else:
		if cellObj.value  not in reqt:
			sheet[cellObj.coordinate].fill = redFill


#validate AT USERID
cursor.execute("select lgb_online_user_id from lgb_employee order by 1;")  
row = cursor.fetchone()
at_user = []  
while row:  
    at_user.append(row[0])    
    row = cursor.fetchone()
for cellObj in sheet.columns[2]:
	if cellObj.value == None:
		pass
	else:
		if cellObj.value  not in at_user:
			sheet[cellObj.coordinate].fill = redFill

#validate REQ USERID
cursor.execute("select lgb_online_user_id from lgb_employee order by 1;")  
row = cursor.fetchone()
req_user = []  
while row:  
    req_user.append(row[0])    
    row = cursor.fetchone()
for cellObj in sheet.columns[3]:
	if cellObj.value == None:
		pass
	else:
		if cellObj.value  not in req_user:
			sheet[cellObj.coordinate].fill = redFill
			
#validate ASSETSTATUS
cursor.execute("select lgv_request_type_desc + lgu_request_subtype_desc from lgb_requesttype join lgb_requestsubtype on lgv_request_type_id = lgu_request_type_id;")  
row = cursor.fetchone()
reqst = []  
while row:  
    reqst.append(row[0])    
    row = cursor.fetchone()
for cellObj in sheet.columns[4]:
	if cellObj.value == None:
		pass
	else:
		if cellObj.value  not in reqst:
			sheet[cellObj.coordinate].fill = redFill


print "\n Saving..."
wb.save('sched_data_val.xlsx')
