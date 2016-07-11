import sys
import pymssql  
import openpyxl

from _profile_helper import HEADER_FN
from _profile_helper import apply
from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import get_column_letter

#Connect to database
database_in = 'SUPPORT' + raw_input('Enter support database number: ')
try:
    conn = pymssql.connect(server='ATLASSPSQL1', user='sa', password='titp4sa', database=database_in)  
except pymssql.Error as e:
	print "An error has occurred: ", e
	sys.exit(1)
cursor = conn.cursor()

#Open the read excel file and the results excel file 
Tk().withdraw()
filename = askopenfilename()
wb = openpyxl.load_workbook(filename, use_iterators=True)
sheet = wb.get_sheet_by_name('Assets')
nwb = openpyxl.Workbook(write_only=True)

row_count = sheet.max_row
print "\n Starting data profiling..."
#Collect header names
header_cells = []
max_col = get_column_letter(sheet.max_column)
range_expr = 'A1:' + max_col + '1'
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		header_cells.append(cell)

#Run the main functions for each header name		
for header in header_cells:
	if HEADER_FN.has_key(header.value):
		header_col = get_column_letter(header.column)
		range_expr = header_col + '2:' + header_col + str(row_count)
		fn = HEADER_FN[header.value]

		results = apply(fn, header.value, sheet, range_expr, cursor)
		
		if results:
			nws = nwb.create_sheet(header.value)
			nws.append([header.value,'Row','Errors'])
			for row in results:
				nws.append(row)
#Save to new file		
print "\n Saving..."
nwb.save('output.xlsx')