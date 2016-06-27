import pymssql  
import openpyxl
from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border

def valB(sheet, row_num):
    row_count = sheet.max_row
    range_expr = 'B' + str(row_num) +':B' + str(row_num)
    for row in sheet.iter_rows(range_string=range_expr):
        for cell in row:
	    if cell.value == None:
	        return cell.value
	    else:
		return cell.value


database_in = 'SUPPORT' + raw_input('Enter support database number: ')
try:
    conn = pymssql.connect(server='ATLASSPSQL1', user='sa', password='titp4sa', database=database_in)  
except pymssql.Error as e:
	print "An error has occurred: ", e.value	
cursor = conn.cursor()  
Tk().withdraw()
filename = askopenfilename()

#wb = openpyxl.load_workbook(filename, use_iterators = True)
wb = openpyxl.load_workbook(filename)
sheet = wb.get_sheet_by_name('Assets')
#nwb = openpyxl.Workbook(write_only=True)

#one sheet at a time - build specifically for assets	
#connect to database	
#open excel file
#function for each column to be validated
#in function: grab db field data and put in list
#iterate through the columns cells, if it fails validation, mark the cell as red, append comment as to why in front


redFill = PatternFill(start_color='FFFF0000',
							end_color='FFFF0000',
							fill_type='solid')
row_count = sheet.max_row

#validate PROPERTYID
cursor.execute("SELECT lgp_external_property_id FROM lgb_property ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
#nws = nwb.create_sheet('ExtPropID')
range_expr = 'A2:A' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		if cell.value == None:
		    cell.fill = redFill
		else:
			if cell.value not in results:
				cell.fill = redFill
				#nws.append([cell.value])
				
#validate ASSETCLASS
cursor.execute("SELECT asc_asset_class_description FROM asm_assetclass ORDER BY 1;")  
row = cursor.fetchone()	
results = []
while row:  
    results.append(row[0])    
    row = cursor.fetchone()		
#nws = nwb.create_sheet('AssetClass')
range_expr = 'B2:B' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		if cell.value == None:
			cell.fill = redFill
		else:
			if cell.value not in results:
			    cell.fill = redFill
				#nws.append([cell.value])

#validate ASSETNAME
range_expr = 'C2:C' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
    for cell in row:
	    if cell.value == None:
		    cell.fill = redFill

#validate ASSETNUMBER	
range_expr = 'D2:D' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
    for cell in row:
	    if cell.value == None:
		    cell.fill = redFill

#validate SERIALNUMBER
range_expr = 'F2:F' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
    for cell in row:
	    if cell.value == None:
		    cell.fill = redFill
			
#validate ASSETRANK
cursor.execute("SELECT asr_asset_rank_description from asm_assetrank ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
#nws = nwb.create_sheet('AssetRank')
range_expr = 'G2:G' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		if cell.value == None:
			cell.fill = redFill
		else:
			if cell.value not in results:
				cell.fill = redFill
				#nws.append([cell.value])

#validate MAKE
range_expr = 'H2:H' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
    for cell in row:
	    if cell.value == None:
		    cell.fill = redFill
			
#validate MODEL
range_expr = 'I2:I' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
    for cell in row:
	    if cell.value == None:
		    cell.fill = redFill				
				
#validate ASSETSTATUS		
cursor.execute("SELECT asg_asset_status_description from asm_assetstatus ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
#nws = nwb.create_sheet('AssetStatus')
range_expr = 'J2:J' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		if cell.value == None:
			cell.fill = redFill
		else:
			if cell.value not in results:
				cell.fill = redFill
				#nws.append([cell.value])
				
#validate ASSETKEYWORD		
cursor.execute("SELECT ask_asset_keyword_name from asm_asset_keyword ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
#nws = nwb.create_sheet('AssetKeyword')
range_expr = 'AL2:AL' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	for cell in row:
		if cell.value == None:
			cell.fill = redFill
		else:
			if cell.value not in results:
				cell.fill = redFill
				#nws.append([cell.value])
				
				
print "\n Saving..."
wb.save('output.xlsm')
#nwb.save('output.xlsx')
