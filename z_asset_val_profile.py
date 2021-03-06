import sys
import pymssql  
import openpyxl

from Tkinter import Tk
from tkFileDialog import askopenfilename
from openpyxl.styles import Color, PatternFill, Font, Border


database_in = 'SUPPORT' + raw_input('Enter support database number: ')
try:
    conn = pymssql.connect(server='ATLASSPSQL1', user='sa', password='titp4sa', database=database_in)  
except pymssql.Error as e:
	print "An error has occurred: ", e
	sys.exit(1)
cursor = conn.cursor()  
Tk().withdraw()
filename = askopenfilename()

wb = openpyxl.load_workbook(filename, use_iterators=True)
sheet = wb.get_sheet_by_name('Assets')
nwb = openpyxl.Workbook(write_only=True)

#one sheet at a time - build specifically for assets	
#connect to database	
#open excel file
#function for each column to be validated
#in function: grab db field data and put in list
#iterate through the columns cells, if it fails validation, mark the cell as red, append comment as to why in front



row_count = sheet.max_row

#validate PROPERTYID
cursor.execute("SELECT lgp_external_property_id FROM lgb_property ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()

nws = nwb.create_sheet('ExtPropID')
row_number = 1
nws.append(['ExtPropID','Row','Errors'])
range_expr = 'A2:A' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
    row_number += 1
    for cell in row:
		if cell.value == None:
			error_msg += "ExtPropertyID required; "
		else:
		    if cell.value not in results:
				error_msg += "Property ID not found; "
		
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])
				
#validate ASSETCLASS
cursor.execute("SELECT asc_asset_class_description FROM asm_assetclass ORDER BY 1;")  
row = cursor.fetchone()	
results = []
while row:  
    results.append(row[0])    
    row = cursor.fetchone()	

nws = nwb.create_sheet('AssetClass')
row_number = 1
nws.append(['AssetClass','Row','Errors'])
range_expr = 'B2:B' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""	
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "AssetClass required; "
		else:
		    if cell.value not in results:
				error_msg += "Asset class not found; "
				
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])

#validate ASSETNAME
nws = nwb.create_sheet('AssetName')
row_number = 1
nws.append(['AssetName','Row','Errors'])	
range_expr = 'C2:C' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Asset name required; "
		
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])

#validate ASSETNUMBER	
nws = nwb.create_sheet('AssetNumber')
row_number = 1
nws.append(['AssetNumber','Row','Errors'])		
range_expr = 'D2:D' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Asset number required; "
			
		if error_msg != "":
		    nws.append([cell.value, row_number, error_msg])

#validate SERIALNUMBER
nws = nwb.create_sheet('SerialNumber')
row_number = 1
nws.append(['SerialNumber','Row','Errors'])
range_expr = 'F2:F' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Serial number required; "
			
		if error_msg != "":
		    nws.append([cell.value, row_number, error_msg])
			
#validate ASSETRANK
cursor.execute("SELECT asr_asset_rank_description from asm_assetrank ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
	
nws = nwb.create_sheet('AssetRank')
row_number = 1
nws.append(['AssetRank','Row','Errors'])
range_expr = 'G2:G' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Asset rank required; "
		else:
			if cell.value not in results:
				error_msg += "Asset rank not found; "
		
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])

#validate MAKE
nws = nwb.create_sheet('Make')
row_number = 1
nws.append(['Make','Row','Errors'])
range_expr = 'H2:H' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Make required; "
			
		if error_msg != "":
		    nws.append([cell.value, row_number, error_msg])
			
#validate MODEL
nws = nwb.create_sheet('Model')
row_number = 1
nws.append(['Model','Row','Errors'])
range_expr = 'I2:I' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Model required; "
		
		if error_msg != "":
		    nws.append([cell.value, row_number, error_msg])			
				
#validate ASSETSTATUS		
cursor.execute("SELECT asg_asset_status_description from asm_assetstatus ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
	
nws = nwb.create_sheet('AssetStatus')
row_number = 1
nws.append(['AssetStatus','Row','Errors'])
range_expr = 'J2:J' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			error_msg += "Asset status required; "
		else:
			if cell.value not in results:
				error_msg += "Asset status not found; "
		
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])	
				
#validate ASSETKEYWORD		
cursor.execute("SELECT ask_asset_keyword_name from asm_asset_keyword ORDER BY 1;")  
row = cursor.fetchone()
results = []  
while row:  
    results.append(row[0])    
    row = cursor.fetchone()
	
nws = nwb.create_sheet('AssetKeyword')
row_number = 1
nws.append(['AssetKeyword','Row','Errors'])
range_expr = 'AL2:AL' + str(row_count)
for row in sheet.iter_rows(range_string=range_expr):
	error_msg = ""
	row_number += 1
	for cell in row:
		if cell.value == None:
			pass
		else: 
			if cell.value not in results:
				error_msg =+ "Asset keyword not found; "
		
		if error_msg != "":
			nws.append([cell.value, row_number, error_msg])	
				
				
print "\n Saving..."
nwb.save('output.xlsx')
	



